#1.导包
import openpyxl
import os
import logging


logger = logging.getLogger(__name__)

# 2.打开excel
# 1.动态获取文件路径
def get_file(filename="case01_add_goods.xlsx"):
    cur_path = os.path.realpath(__file__)
    print("当前路径为%s" %cur_path)
    filename = os.path.dirname(os.path.dirname(cur_path)) + "\Data\\" + filename
    print("要打开的文件为%s" %filename)
    # 2.打开文件，获取workbook对象
    workbook = openpyxl.load_workbook(filename)
    return workbook
# 3.获取sheet对象

workbook = get_file()
sheet = workbook["无需登录添加商品"]
print("aaaaaaaaaa")
logger.info("这里是打印的logging信息")
# 4.动态获取sheet对象
sheets = workbook.sheetnames
print(sheets)
# sheet = sheets[0]

# 5.获取数据的行数
print("数据行数为：%s" %sheet.max_row)

# 6.获取数据的列数
print("数据列数为：%s" %sheet.max_column)

#7.读取3行9列的数据
data = sheet.cell(3,9).value

print("workbook的值为%s" %workbook)
print("数据的行数为%s" %sheet.max_row)
print("数据的列数为%s" %sheet.max_column)
print("3行9列的数据为%s" %data)