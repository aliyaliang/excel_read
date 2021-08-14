from config import base_path
import os
import openpyxl
from config import cell_config
import json


class FileTool:
    # 1.初始化
    def __init__(self, filename):
        # 1.动态获取文件路径
        # self.filename = base_path + os.sep + "Data" + os.sep + filename
        self.cur_path = os.path.realpath(__file__)
        print("当前路径为%s" %self.cur_path)
        self.filename = os.path.dirname(os.path.dirname(self.cur_path)) + "\data\\" + filename
        print("要打开的文件为%s" %self.filename)
        # 2.打开文件，获取workbook对象
        self.workbook = openpyxl.load_workbook(self.filename)
        # 3.获取sheet表单对象，动态打开第一个sheet
        self.sheet = self.workbook[self.workbook.sheetnames[0]]

        # 4.获取总行数（读取数据，遍历使用）
        self.row = self.sheet.max_row
        print("总行数为：",self.row)
        # self.filename = filename  # 新增

    # 2.读取Excel
    def read_excel(self):
        # 1.新建空列表，存储每行数据
        case = list()
        # 2.遍历每行数据
        for i in range(2, self.row+1):
            # 新建空字典，存储每行数据
            data = dict()
            # 判断是否执行
            if self.sheet.cell(i, cell_config.get("is_run")).value == "是":
                try:
                    # 读取数据，追加到字典
                    data["path"] = self.sheet.cell(i, cell_config.get("path")).value
                    data["method"] = self.sheet.cell(i, cell_config.get("method")).value.lower()
                    data["headers"] = self.sheet.cell(i, cell_config.get("headers")).value
                    data["param_type"] = self.sheet.cell(i, cell_config.get("param_type")).value
                    data["params"] = self.sheet.cell(i, cell_config.get("params")).value
                    data["expect"] = self.sheet.cell(i, cell_config.get("expect")).value
                    data["param_type"] = self.sheet.cell(i, cell_config.get("param_type")).value
                    # 记录每行数据的行与列，执行完写入结果使用：
                    data["x_y"] = [i, cell_config.get("result")]
                    # 将字典追加到空列表中
                    case.append(data)
                    # 将读取结果写入excel中
                    self.write_excel([i, cell_config.get("desc")], "数据读取完成~")
                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], e)

        # 3.将列表数据写入json
        self.write_json(case, "case.json")
        print("读取的数据为：", case)

    # 3.写入Excel
    def write_excel(self, x_y, msg):
        try:
            # x_y参数格式为列表，如：[2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)

    # 4.读取json
    def read_json(self, file_name="case.json"):
        file_name = os.path.dirname(os.path.dirname(self.cur_path)) + "\data\\" + file_name
        with open(file_name, "r", encoding="utf-8") as fp:
            return json.load(fp)

    # 5.写入json
    def write_json(self, case, file_name):
        file_name = os.path.dirname(os.path.dirname(self.cur_path)) + "\data\\" + file_name
        with open(file_name, "w", encoding="utf-8") as fp:
            json.dump(case, fp, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    # FileTool = FileTool("case01_add_goods.xlsx")
    # # 调用FileTool class下面read_excel方法下面的case
    # print(FileTool.read_excel())
    # case = FileTool.read_excel().case
    # FileTool.write_excel(case.get("x_y"))
    # FileTool.read_excel()
    # FileTool.read_json("case01_add_goods.xlsx")
    # 1.调用FileTool类的初始化方法：
    FileTool("case01_add_goods.xlsx")
    # 2.调用FileTool类下面read_excel方法：
    print(FileTool.read_excel)




